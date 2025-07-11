import subprocess
import sys
import time
import random
import logging

# Initial warning and confirmation
print("WARNING: This script matches headers case-insensitively (e.g., zip, Zip, ZIP),")
print("but special characters or extra whitespace in header names can prevent detection.")
response = input("Do you want to continue? (y/n): ").strip().lower()
if response not in ('y', 'yes'):
    print("Aborting script as requested.")
    sys.exit(0)

# Ask user for the full file path of the Excel workbook
input_file = input("Enter the full path to your Excel file (including extension): ").strip()
if not input_file:
    print("No file path provided. Exiting.")
    sys.exit(1)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s')

def install_package(pkg):
    """Install a package via pip if missing."""
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg], stdout=subprocess.DEVNULL)

# Ensure required packages
for pkg in ("openpyxl", "geopy", "requests"):
    try:
        __import__(pkg)
    except ImportError:
        logging.info(f"Installing missing package: {pkg}")
        install_package(pkg)

import openpyxl
from openpyxl.styles import PatternFill
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderUnavailable
import requests
from requests.exceptions import RequestException

# Generate output filename with epoch timestamp
timestamp = int(time.time())
output_file = f"output_{timestamp}.xlsx"

# Load workbook
try:
    wb = openpyxl.load_workbook(input_file)
except FileNotFoundError:
    logging.error(f"File not found: {input_file}")
    sys.exit(1)
ws = wb.active

# Map headers to columns (case-insensitive, stripped)
headers = {}
for idx, cell in enumerate(ws[1], start=1):
    if cell.value is not None:
        header = str(cell.value).strip().lower()
        headers[header] = idx

col_zip     = headers.get("zip")
col_address = headers.get("full_address")

if not col_zip or not col_address:
    available = ", ".join(f"'{h}'" for h in headers.keys())
    logging.error(f"Could not find required headers. Available: {available}")
    sys.exit(1)

# Initialize Nominatim
geolocator = Nominatim(user_agent="zip_enricher", timeout=10)
yellow_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")

def lookup_postcode(address, retries=3):
    """Attempt to geocode the address and return the postcode, with retries."""
    for attempt in range(1, retries + 1):
        try:
            loc = geolocator.geocode(address, addressdetails=True)
            if loc and 'postcode' in loc.raw.get('address', {}):
                return loc.raw['address']['postcode']
            return ""
        except (GeocoderTimedOut, GeocoderUnavailable, RequestException) as e:
            wait = min(20, 2 ** attempt)
            logging.warning(f"Attempt {attempt} failed for '{address}': {e}. Retrying in {wait}s.")
            time.sleep(wait)
    logging.error(f"All retries failed for address: '{address}'.")
    return ""

# Iterate through rows and fill blank zips
for row in range(2, ws.max_row + 1):
    zip_cell = ws.cell(row=row, column=col_zip)
    if not zip_cell.value or str(zip_cell.value).strip() == "":
        address = ws.cell(row=row, column=col_address).value or ""
        postcode = lookup_postcode(address)
        if postcode:
            zip_cell.value = postcode
            zip_cell.fill = yellow_fill
            logging.info(f"Row {row}: filled ZIP {postcode}")
        else:
            logging.info(f"Row {row}: no ZIP found for '{address}'")
        time.sleep(random.uniform(1, 5))  # rate-limit

# Trim whitespace in all cells
for r in ws.iter_rows(min_row=2):
    for c in r:
        if isinstance(c.value, str):
            c.value = c.value.strip()

# Save updated workbook
wb.save(output_file)
logging.info(f"Done. Updated file saved as {output_file}")
